from openpyxl import load_workbook
from .util import create_ticker


def get_stock_dcf_analysis(symbol: str):
    """
    Generate a comprehensive Discounted Cash Flow (DCF) valuation Excel spreadsheet for a given stock.

    The DCF method estimates a company's intrinsic value by projecting future free cash flows
    and discounting them back to present value using the Weighted Average Cost of Capital (WACC).
    This tool generates a professional Excel file with all calculations, growth assumptions,
    discount rates, and fair value estimates.

    Args:
        symbol (str):
            Stock ticker symbol, e.g., "AAPL", "TSLA", "QCOM" (case-insensitive).

    Returns:
        dict: {
            "symbol": str,                      # Stock ticker symbol
            "file_path": str,                   # Absolute path to the generated Excel file (local)
            "description": str,                 # Brief description of the analysis
            "dcf_value": dict,                  # DCF valuation results
            "discount_rate": dict,              # Discount rate components (WACC, Cost of Equity, Cost of Debt)
            "growth_estimates": dict,           # Historical growth rates (Revenue, FCF, EBITDA, Net Income)
            "investment_recommendation": str    # Buy/Hold/Sell recommendation
        }

    What's in the Excel file:
        - Discount Rate Estimates (WACC, Cost of Equity, Cost of Debt)
        - Growth Estimates (Historical 3-year CAGR for Revenue, FCF, EBITDA, Net Income)
        - DCF Template (10-year cash flow projections with present value calculations)
        - DCF Valuation (Enterprise Value, Fair Price, Current Price, Margin of Safety)
        - Investment Recommendation (Buy/Sell based on fair value vs current price)

    Important:
        The Excel file is fully editable. Users can adjust growth assumptions, discount rates,
        and other parameters to perform sensitivity analysis. All formulas will automatically recalculate.

    Example:
        >>> get_stock_dcf_analysis("AAPL")
        {
            "symbol": "AAPL",
            "file_path": "/tmp/defeatbeta/AAPL.xlsx",
            "description": "DCF Valuation Analysis for AAPL",
            "dcf_value": {
                "fair_price": 185.50,
                "current_price": 150.25,
                "enterprise_value": 2500000000000,
                "equity_value": 2450000000000
            },
            "discount_rate": {
                "wacc": 0.089,
                "cost_of_equity": 0.095,
                "cost_of_debt": 0.032
            },
            "growth_estimates": {
                "revenue_cagr_3y": 0.085,
                "fcf_cagr_3y": 0.12
            },
            "investment_recommendation": "Buy (Fair value $185.50 is 23.5% higher than current $150.25)",
            "note": "DCF Excel file generated at: /tmp/defeatbeta/AAPL.xlsx..."
        }
    """

    symbol = symbol.upper()
    ticker = create_ticker(symbol)

    try:
        result = ticker.dcf()

        if not result or "file_path" not in result:
            return {
                "symbol": symbol,
                "error": "Failed to generate DCF analysis. The ticker may not have sufficient financial data."
            }

        file_path = result["file_path"]

        # Read the Excel file and extract key data
        # Use data_only=True to read calculated values instead of formulas
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Extract DCF Value section (typically in top right)
        # These positions may need adjustment based on actual Excel layout
        dcf_value = {}
        discount_rate = {}
        growth_estimates = {}

        # Helper function to safely get cell value
        def get_cell_value(cell):
            value = cell.value
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return value
            return str(value)

        # Try to extract key metrics by scanning the sheet
        for row in ws.iter_rows():
            for cell in row:
                cell_value = get_cell_value(cell)
                if cell_value is None:
                    continue

                cell_str = str(cell_value).strip()

                # Extract DCF valuation metrics
                if "Fair Price" in cell_str or "Fair Value" in cell_str:
                    # Fair price is usually in the next cell
                    fair_price_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    dcf_value["fair_price"] = get_cell_value(fair_price_cell)

                elif "Current Price" in cell_str or "Market Price" in cell_str:
                    current_price_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    dcf_value["current_price"] = get_cell_value(current_price_cell)

                elif "Enterprise Value" in cell_str and "Total" not in cell_str:
                    ev_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    dcf_value["enterprise_value"] = get_cell_value(ev_cell)

                elif "Equity Value" in cell_str:
                    equity_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    dcf_value["equity_value"] = get_cell_value(equity_cell)

                elif cell_str == "WACC":
                    wacc_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    discount_rate["wacc"] = get_cell_value(wacc_cell)

                elif "Cost of Equity" in cell_str:
                    coe_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    discount_rate["cost_of_equity"] = get_cell_value(coe_cell)

                elif "Cost of Debt" in cell_str:
                    cod_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    discount_rate["cost_of_debt"] = get_cell_value(cod_cell)

                elif "Revenue" in cell_str and "CAGR" in cell_str:
                    rev_cagr_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    growth_estimates["revenue_cagr_3y"] = get_cell_value(rev_cagr_cell)

                elif "FCF" in cell_str and "CAGR" in cell_str:
                    fcf_cagr_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    growth_estimates["fcf_cagr_3y"] = get_cell_value(fcf_cagr_cell)

                elif ("Buy" in cell_str or "Sell" in cell_str or "Hold" in cell_str) and len(cell_str) < 20:
                    recommendation = cell_str

        # Determine investment recommendation based on fair price vs current price
        recommendation = "N/A"
        if dcf_value.get("fair_price") and dcf_value.get("current_price"):
            try:
                fair = float(dcf_value["fair_price"])
                current = float(dcf_value["current_price"])
                margin = ((fair - current) / current) * 100

                if margin > 20:
                    recommendation = f"Buy (Fair value ${fair:.2f} is {margin:.1f}% higher than current ${current:.2f})"
                elif margin < -20:
                    recommendation = f"Sell (Fair value ${fair:.2f} is {abs(margin):.1f}% lower than current ${current:.2f})"
                else:
                    recommendation = f"Hold (Fair value ${fair:.2f} is close to current ${current:.2f}, margin: {margin:.1f}%)"
            except (ValueError, TypeError):
                pass

        wb.close()

        return {
            "symbol": symbol,
            "file_path": file_path,
            "description": result.get("description", f"DCF Valuation Analysis for {symbol}"),
            "dcf_value": dcf_value,
            "discount_rate": discount_rate,
            "growth_estimates": growth_estimates,
            "investment_recommendation": recommendation,
            "note": (
                f"DCF Excel file generated at: {file_path}\n"
                "The Excel file is fully editable for sensitivity analysis. "
                "Key valuation metrics have been extracted above."
            )
        }

    except Exception as e:
        return {
            "symbol": symbol,
            "error": f"Error generating DCF analysis: {str(e)}"
        }
