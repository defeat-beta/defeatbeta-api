"""Build a formula-driven DCF Excel workbook from MCP `dcf_data` JSON.

Reads the JSON payload produced by the defeatbeta MCP server's
`get_stock_dcf_analysis` tool and writes a single-sheet .xlsx where every
projection, NPV, and fair-price cell is an Excel formula referencing
input cells. The user can edit any assumption (discount rate, growth
rates, cash, share count, current price) and the model recomputes.

Usage:
    python build_dcf_excel.py <input.json> [output.xlsx]

Defaults output to ./{SYMBOL}_DCF.xlsx in the current working directory.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ===== Palette (3 blues + 1 grey + white, per 3-statement-model conventions) =====
SECTION_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
COLUMN_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")   # light blue
INPUT_FILL = PatternFill("solid", fgColor="F2F2F2")           # light grey
KEY_TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")       # medium blue

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BLACK_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
BLUE_INPUT_FONT = Font(name="Calibri", size=11, color="0000FF")
BLACK_FORMULA_FONT = Font(name="Calibri", size=11, color="000000")

THIN_GREY = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ===== Number formats =====
FMT_INT = "#,##0"
FMT_PCT = "0.00%"
FMT_DEC2 = "0.00"
FMT_FLOAT4 = "0.0000"


@dataclass
class CellRef:
    """A typed wrapper around an Excel cell address like 'C4'."""

    column: str
    row: int

    @property
    def addr(self) -> str:
        return f"{self.column}{self.row}"

    def __str__(self) -> str:  # so f-strings can embed it directly
        return self.addr


@dataclass
class BuildContext:
    """Tracks cursor position and named cell refs as we write sections."""

    ws: Worksheet
    row: int = 1
    refs: Dict[str, CellRef] = field(default_factory=dict)

    def advance(self, n: int = 1) -> None:
        self.row += n

    def remember(self, name: str, col: str, row: int) -> CellRef:
        ref = CellRef(col, row)
        self.refs[name] = ref
        return ref

    def get(self, name: str) -> CellRef:
        return self.refs[name]


# ===== Cell writers =====
def write_section_header(ctx: BuildContext, title: str, span_cols: int = 7) -> None:
    """Write a section banner spanning N columns."""
    start = "B"
    end = get_column_letter(1 + span_cols)  # B=2, so end col index = 2 + span_cols - 1
    cell = ctx.ws[f"{start}{ctx.row}"]
    cell.value = title
    cell.font = WHITE_BOLD
    cell.fill = SECTION_HEADER_FILL
    cell.alignment = LEFT_ALIGN
    ctx.ws.merge_cells(f"{start}{ctx.row}:{end}{ctx.row}")
    ctx.ws.row_dimensions[ctx.row].height = 22
    ctx.advance()


def write_label(ctx: BuildContext, col: str, label: str, bold: bool = False) -> None:
    cell = ctx.ws[f"{col}{ctx.row}"]
    cell.value = label
    cell.font = BLACK_BOLD if bold else Font(name="Calibri", size=11)
    cell.alignment = LEFT_ALIGN


def write_column_header(ctx: BuildContext, col: str, label: str) -> None:
    cell = ctx.ws[f"{col}{ctx.row}"]
    cell.value = label
    cell.font = BLACK_BOLD
    cell.fill = COLUMN_HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def write_input(ctx: BuildContext, col: str, value, number_format: str = FMT_INT) -> CellRef:
    cell = ctx.ws[f"{col}{ctx.row}"]
    cell.value = value
    cell.font = BLUE_INPUT_FONT
    cell.fill = INPUT_FILL
    cell.number_format = number_format
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER
    return CellRef(col, ctx.row)


def write_formula(
    ctx: BuildContext,
    col: str,
    formula: str,
    number_format: str = FMT_INT,
    is_key_total: bool = False,
) -> CellRef:
    cell = ctx.ws[f"{col}{ctx.row}"]
    cell.value = formula
    cell.font = BLACK_BOLD if is_key_total else BLACK_FORMULA_FONT
    if is_key_total:
        cell.fill = KEY_TOTAL_FILL
    cell.number_format = number_format
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER
    return CellRef(col, ctx.row)


# ===== Section builders =====
def build_discount_rate_section(ctx: BuildContext, dr: dict) -> None:
    """Section 1: WACC inputs & derivations.

    Layout:
        B = input label, C = input value
        E = derived label, F = derived value
    """
    write_section_header(ctx, f"DISCOUNT RATE ESTIMATES ({dr.get('report_date', 'N/A')})")
    ctx.advance()  # blank spacer row

    # Column headers (B/C inputs | E/F derived)
    write_column_header(ctx, "B", "Input")
    write_column_header(ctx, "C", "Value")
    write_column_header(ctx, "E", "Derived")
    write_column_header(ctx, "F", "Value")
    ctx.advance()

    # --- Inputs (B/C) ---
    inputs_start = ctx.row
    write_label(ctx, "B", "Market Cap (USD)")
    market_cap = write_input(ctx, "C", dr["market_cap"], FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Beta (5Y)")
    beta = write_input(ctx, "C", dr["beta_5y"], FMT_DEC2)
    ctx.advance()

    write_label(ctx, "B", "Total Debt (USD)")
    total_debt = write_input(ctx, "C", dr["total_debt"], FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Interest Expense (USD)")
    interest_expense = write_input(ctx, "C", dr["interest_expense"], FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Risk-Free Rate (10Y Treasury)")
    risk_free = write_input(ctx, "C", dr["risk_free_rate"], FMT_PCT)
    ctx.advance()

    write_label(ctx, "B", "Expected Market Return (S&P 500)")
    market_return = write_input(ctx, "C", dr["expected_market_return"], FMT_PCT)
    ctx.advance()

    # --- Derived (E/F), starting at inputs_start ---
    derived_row = inputs_start
    write_label_at(ctx, "E", derived_row, "Weight of Debt")
    weight_of_debt = write_formula_at(
        ctx, "F", derived_row, f"={total_debt}/({market_cap}+{total_debt})", FMT_PCT
    )

    derived_row += 1
    write_label_at(ctx, "E", derived_row, "Weight of Equity")
    weight_of_equity = write_formula_at(
        ctx, "F", derived_row, f"={market_cap}/({market_cap}+{total_debt})", FMT_PCT
    )

    derived_row += 1
    write_label_at(ctx, "E", derived_row, "Cost of Debt")
    cost_of_debt = write_formula_at(
        ctx, "F", derived_row, f"={interest_expense}/{total_debt}", FMT_PCT
    )

    derived_row += 1
    write_label_at(ctx, "E", derived_row, "Cost of Equity")
    cost_of_equity = write_formula_at(
        ctx,
        "F",
        derived_row,
        f"={risk_free}+{beta}*({market_return}-{risk_free})",
        FMT_PCT,
    )

    derived_row += 1
    write_label_at(ctx, "E", derived_row, "Tax Rate")
    tax_rate = write_input_at(ctx, "F", derived_row, dr["tax_rate"], FMT_PCT)

    derived_row += 1
    write_label_at(ctx, "E", derived_row, "WACC", bold=True)
    wacc = write_formula_at(
        ctx,
        "F",
        derived_row,
        f"={weight_of_debt}*{cost_of_debt}*(1-{tax_rate})+{weight_of_equity}*{cost_of_equity}",
        FMT_PCT,
        is_key_total=True,
    )

    # Save important refs
    ctx.refs["wacc"] = wacc
    ctx.refs["total_debt_input"] = total_debt
    ctx.refs["risk_free_input"] = risk_free
    # Ensure cursor is below the wider of the two columns
    ctx.row = max(ctx.row, derived_row + 1)
    ctx.advance()  # blank spacer


def write_label_at(ctx: BuildContext, col: str, row: int, label: str, bold: bool = False) -> None:
    cell = ctx.ws[f"{col}{row}"]
    cell.value = label
    cell.font = BLACK_BOLD if bold else Font(name="Calibri", size=11)
    cell.alignment = LEFT_ALIGN


def write_input_at(ctx: BuildContext, col: str, row: int, value, number_format: str) -> CellRef:
    cell = ctx.ws[f"{col}{row}"]
    cell.value = value
    cell.font = BLUE_INPUT_FONT
    cell.fill = INPUT_FILL
    cell.number_format = number_format
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER
    return CellRef(col, row)


def write_formula_at(
    ctx: BuildContext,
    col: str,
    row: int,
    formula: str,
    number_format: str,
    is_key_total: bool = False,
) -> CellRef:
    cell = ctx.ws[f"{col}{row}"]
    cell.value = formula
    cell.font = BLACK_BOLD if is_key_total else BLACK_FORMULA_FONT
    if is_key_total:
        cell.fill = KEY_TOTAL_FILL
    cell.number_format = number_format
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER
    return CellRef(col, row)


def build_growth_estimates_section(ctx: BuildContext, ge: dict) -> None:
    """Section 2: Revenue 3Y CAGR, EPS multi-year CAGR, Treasury 5Y avg."""
    write_section_header(ctx, "GROWTH ESTIMATES")
    ctx.advance()

    currency = ge.get("currency", "USD")

    # ---- Revenue block ----
    write_label(ctx, "B", f"Revenue ({currency}) — 3Y Historical", bold=True)
    ctx.advance()

    write_column_header(ctx, "B", "Date")
    write_column_header(ctx, "C", "Value")
    write_column_header(ctx, "D", "YoY Growth")
    ctx.advance()

    revenue_details: List[dict] = ge["revenue"]["details"]
    revenue_value_cells: List[CellRef] = []
    for item in revenue_details:
        write_label(ctx, "B", str(item.get("date", "N/A")))
        val_ref = write_input(ctx, "C", item["value"], FMT_INT)
        write_input(ctx, "D", item["yoy"], FMT_PCT)
        revenue_value_cells.append(val_ref)
        ctx.advance()

    # Revenue 3Y CAGR formula (preserve dcf_data semantics: Turned Positive/Negative)
    write_label(ctx, "B", "Revenue 3Y CAGR", bold=True)
    if len(revenue_value_cells) >= 3:
        start = revenue_value_cells[0]
        end = revenue_value_cells[2]
        formula = (
            f'=IF({start}<=0,IF({end}>0,"Turned Positive","N/A"),'
            f'IF({end}<=0,"Turned Negative",POWER({end}/{start},1/2)-1))'
        )
        revenue_cagr = write_formula(ctx, "C", formula, FMT_PCT)
    else:
        revenue_cagr = write_input(ctx, "C", ge["revenue"].get("cagr_3y", 0) or 0, FMT_PCT)
    ctx.refs["revenue_cagr"] = revenue_cagr
    ctx.advance(2)

    # ---- EPS block ----
    write_label(ctx, "B", f"EPS TTM ({currency}) — Annual Snapshots", bold=True)
    ctx.advance()

    write_column_header(ctx, "B", "Date")
    write_column_header(ctx, "C", "EPS")
    write_column_header(ctx, "D", "YoY Growth")
    ctx.advance()

    eps_details: List[dict] = ge.get("eps", {}).get("details", [])
    eps_value_cells: List[CellRef] = []
    for item in eps_details:
        write_label(ctx, "B", str(item.get("date", "N/A")))
        val_ref = write_input(ctx, "C", item["value"], FMT_DEC2)
        write_input(ctx, "D", item["yoy"], FMT_PCT)
        eps_value_cells.append(val_ref)
        ctx.advance()

    # EPS CAGR — pick the earliest positive year as the start; if last EPS <=0, fall back to 0
    eps_cagr_years = ge.get("eps", {}).get("cagr_years", 0) or 0
    cagr_label = f"EPS {eps_cagr_years}Y CAGR" if eps_cagr_years > 0 else "EPS CAGR"
    write_label(ctx, "B", cagr_label, bold=True)
    if len(eps_value_cells) >= 2:
        last = eps_value_cells[-1]
        n = len(eps_value_cells)
        inner = "0"
        # Walk from year n-2 down to 0; first positive start wins.
        for j in range(n - 2, -1, -1):
            start = eps_value_cells[j]
            years = n - 1 - j
            inner = f"IF({start}>0,POWER({last}/{start},1/{years})-1,{inner})"
        formula = f"=IF({last}<=0,0,{inner})"
        eps_cagr = write_formula(ctx, "C", formula, FMT_PCT)
    else:
        eps_cagr = write_input(ctx, "C", ge.get("eps", {}).get("cagr_10y", 0) or 0, FMT_PCT)
    ctx.refs["eps_cagr"] = eps_cagr
    ctx.advance(2)

    # ---- Treasury block ----
    write_label(ctx, "B", "US 10Y Treasury Yield — Annual Averages", bold=True)
    ctx.advance()

    write_column_header(ctx, "B", "Year")
    write_column_header(ctx, "C", "Avg Yield")
    ctx.advance()

    treasury_details: List[dict] = ge.get("treasury", {}).get("details", [])
    treasury_value_cells: List[CellRef] = []
    for item in treasury_details:
        write_label(ctx, "B", item.get("year"))
        val_ref = write_input(ctx, "C", item["avg_yield"], FMT_PCT)
        treasury_value_cells.append(val_ref)
        ctx.advance()

    write_label(ctx, "B", "5Y Treasury Avg", bold=True)
    if treasury_value_cells:
        first = treasury_value_cells[0]
        last = treasury_value_cells[-1]
        treasury_avg = write_formula(ctx, "C", f"=AVERAGE({first}:{last})", FMT_PCT)
    else:
        treasury_avg = write_input(
            ctx, "C", ge.get("treasury", {}).get("avg_5y", 0) or 0, FMT_PCT
        )
    ctx.refs["treasury_avg"] = treasury_avg
    ctx.advance(2)


def build_dcf_template_section(ctx: BuildContext, dt: dict) -> None:
    """Section 3: Assumptions, 10-year projections, historical FCF margins."""
    write_section_header(ctx, "DCF TEMPLATE", span_cols=13)
    ctx.advance()

    # ---- Assumptions block (B/C two-column) ----
    write_column_header(ctx, "B", "Assumption")
    write_column_header(ctx, "C", "Value")
    ctx.advance()

    eps_cagr = ctx.refs["eps_cagr"]
    revenue_cagr = ctx.refs["revenue_cagr"]
    treasury_avg = ctx.refs["treasury_avg"]
    wacc = ctx.refs["wacc"]

    # Growth Rate (1~5Y): cap 20%, floor 5%, off raw EPS CAGR
    write_label(ctx, "B", "Future Growth Rate (1~5Y)", bold=True)
    growth_1_5y = write_formula(
        ctx, "C", f"=MIN(MAX({eps_cagr},0.05),0.20)", FMT_PCT, is_key_total=True
    )
    ctx.advance()

    # Growth Rate Terminal (= treasury 5Y avg), written before 6-10Y so the
    # 6-10Y formula can reference it.
    growth_terminal_row = ctx.row + 1  # placeholder; we'll write 6-10Y now and terminal next
    # Write 6-10Y first then terminal — formula uses growth_terminal cell address; we know
    # terminal is at the next row.
    write_label(ctx, "B", "Future Growth Rate (6~10Y)", bold=True)
    # 6-10Y start at year 6: g1 - (g1 - gt)/5
    growth_6_10y = write_formula(
        ctx,
        "C",
        f"={growth_1_5y}-({growth_1_5y}-C{growth_terminal_row})/5",
        FMT_PCT,
    )
    ctx.advance()

    write_label(ctx, "B", "Future Growth Rate (Terminal)", bold=True)
    growth_terminal = write_formula(ctx, "C", f"={treasury_avg}", FMT_PCT, is_key_total=True)
    assert growth_terminal.row == growth_terminal_row, "growth_terminal row mismatch"
    ctx.advance()

    write_label(ctx, "B", "Discount Rate (Default: WACC)", bold=True)
    discount_rate = write_formula(ctx, "C", f"={wacc}", FMT_PCT, is_key_total=True)
    ctx.advance()

    ttm_revenue_label = dt.get("ttm_revenue_label") or "TTM Revenue"
    write_label(ctx, "B", ttm_revenue_label)
    ttm_revenue = write_input(ctx, "C", dt["ttm_revenue"], FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Base FCF (TTM)")
    base_fcf = write_input(ctx, "C", dt["base_fcf"], FMT_INT)
    ctx.advance()

    # Revenue growth rates — used in FCF margin denominator
    write_label(ctx, "B", "Future Revenue Growth (1~5Y)", bold=True)
    rev_growth_1_5y = write_formula(
        ctx, "C", f"=MIN(MAX({revenue_cagr},0.05),0.20)", FMT_PCT
    )
    ctx.advance()

    write_label(ctx, "B", "Future Revenue Growth (6~10Y)", bold=True)
    rev_growth_6_10y = write_formula(
        ctx, "C", f"={rev_growth_1_5y}-({rev_growth_1_5y}-{growth_terminal})/5", FMT_PCT
    )
    ctx.advance(2)  # blank spacer

    # ---- Projections grid: B=label, C=year 0 (TTM), D-M = year 1..10 ----
    proj_cols = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    projections: List[dict] = dt["projections"]

    # Year header row
    write_label(ctx, "B", "Year", bold=True)
    write_column_header(ctx, "C", "Year 0 (TTM)")
    for i in range(1, 11):
        write_column_header(ctx, proj_cols[i], f"Year {i}")
    ctx.advance()

    # Date row
    write_label(ctx, "B", "Date", bold=True)
    for i, col in enumerate(proj_cols):
        date_val = projections[i].get("date", "")
        cell = ctx.ws[f"{col}{ctx.row}"]
        cell.value = str(date_val)
        cell.font = Font(name="Calibri", size=10, italic=True)
        cell.alignment = CENTER_ALIGN
    ctx.advance()

    # FCF row
    fcf_row = ctx.row
    write_label(ctx, "B", "Free Cash Flow", bold=True)
    write_formula(ctx, "C", f"={base_fcf}", FMT_INT)  # year 0 = TTM base FCF
    # Years 1-5: prev * (1 + growth_1_5y)
    for i in range(1, 6):
        prev_col = proj_cols[i - 1]
        write_formula(
            ctx, proj_cols[i], f"={prev_col}{fcf_row}*(1+{growth_1_5y})", FMT_INT
        )
    # Years 6-10: prev * (1 + interpolated rate where rate_i = g1 - (i-5)*(g1-gt)/5)
    for k in range(1, 6):
        i = 5 + k
        prev_col = proj_cols[i - 1]
        rate = f"({growth_1_5y}-{k}*({growth_1_5y}-{growth_terminal})/5)"
        write_formula(
            ctx, proj_cols[i], f"={prev_col}{fcf_row}*(1+{rate})", FMT_INT
        )
    ctx.advance()

    # Terminal Value row — non-zero only at year 10 (last column)
    tv_row = ctx.row
    write_label(ctx, "B", "Terminal Value", bold=True)
    for col in proj_cols[:-1]:
        write_formula(ctx, col, "=0", FMT_INT)
    last_col = proj_cols[-1]
    write_formula(
        ctx,
        last_col,
        f"={last_col}{fcf_row}*(1+{growth_terminal})/({discount_rate}-{growth_terminal})",
        FMT_INT,
    )
    ctx.advance()

    # Total Value row
    total_row = ctx.row
    write_label(ctx, "B", "Total Value", bold=True)
    for col in proj_cols:
        write_formula(ctx, col, f"={col}{fcf_row}+{col}{tv_row}", FMT_INT)
    ctx.refs["total_value_row"] = CellRef("D", total_row)  # used to build NPV range
    ctx.refs["total_value_last_col"] = last_col
    ctx.advance()

    # FCF Margin row
    # year 0: base_fcf / ttm_revenue
    # year 1..5: FCF_i / (TTM_rev * (1+r1)^i)
    # year 6..10: FCF_i / (TTM_rev * (1+r1)^5 * Π(1 + r1 - j*(r1-rt)/5) for j=1..k)
    write_label(ctx, "B", "FCF Margin", bold=True)
    write_formula(ctx, "C", f"={base_fcf}/{ttm_revenue}", FMT_PCT)
    r1 = rev_growth_1_5y
    rt = growth_terminal
    for i in range(1, 6):
        write_formula(
            ctx,
            proj_cols[i],
            f"={proj_cols[i]}{fcf_row}/({ttm_revenue}*POWER(1+{r1},{i}))",
            FMT_PCT,
        )
    for k in range(1, 6):
        i = 5 + k
        base = f"{ttm_revenue}*POWER(1+{r1},5)"
        factors = "*".join(f"(1+{r1}-{j}*({r1}-{rt})/5)" for j in range(1, k + 1))
        write_formula(
            ctx,
            proj_cols[i],
            f"={proj_cols[i]}{fcf_row}/({base}*{factors})",
            FMT_PCT,
        )
    ctx.advance(2)

    # ---- Historical FCF Margin block ----
    historical: List[dict] = dt.get("historical_fcf_margin", []) or []
    if historical:
        write_label(ctx, "B", "Historical FCF Margin (Reference)", bold=True)
        ctx.advance()

        write_label(ctx, "B", "Year", bold=True)
        for i, item in enumerate(historical[:10]):
            col = proj_cols[1 + i]  # start from D, skip C (was year 0 header)
            cell = ctx.ws[f"{col}{ctx.row}"]
            cell.value = str(item.get("date", ""))
            cell.font = BLACK_BOLD
            cell.fill = COLUMN_HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        ctx.advance()

        write_label(ctx, "B", "FCF Margin", bold=True)
        for i, item in enumerate(historical[:10]):
            col = proj_cols[1 + i]
            write_input(ctx, col, item.get("margin"), FMT_PCT)
        ctx.advance(2)

    # Save key cells used by valuation section
    ctx.refs["discount_rate_cell"] = discount_rate
    ctx.refs["total_row"] = CellRef("B", total_row)  # just to recover the row int


def build_dcf_value_section(ctx: BuildContext, dv: dict) -> None:
    """Section 4: Enterprise value, fair price, MoS, Buy/Sell."""
    write_section_header(ctx, f"DCF VALUATION ({dv.get('report_date', 'N/A')})")
    ctx.advance()

    write_column_header(ctx, "B", "Component")
    write_column_header(ctx, "C", "Value")
    ctx.advance()

    discount_rate = ctx.refs["discount_rate_cell"]
    total_row_int = ctx.refs["total_row"].row
    total_last_col = ctx.refs["total_value_last_col"]
    # NPV range: years 1..10 only (D..total_last_col, since C is year 0/TTM)
    npv_range = f"D{total_row_int}:{total_last_col}{total_row_int}"

    write_label(ctx, "B", "Enterprise Value (USD)")
    ev = write_formula(ctx, "C", f"=NPV({discount_rate},{npv_range})", FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Cash & ST Investments (USD)")
    cash = write_input(ctx, "C", dv.get("cash", 0) or 0, FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Total Debt (USD)")
    debt = write_formula(ctx, "C", f"={ctx.refs['total_debt_input']}", FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Equity Value (USD)", bold=True)
    equity = write_formula(ctx, "C", f"={ev}+{cash}-{debt}", FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Outstanding Shares")
    shares = write_input(ctx, "C", dv.get("shares_outstanding", 0) or 0, FMT_INT)
    ctx.advance()

    write_label(ctx, "B", "Fair Price (USD)", bold=True)
    fair_price = write_formula(ctx, "C", f"={equity}/{shares}", FMT_DEC2, is_key_total=True)
    ctx.advance()

    write_label(ctx, "B", "Current Price (USD)")
    current_price = write_input(ctx, "C", dv.get("current_price", 0) or 0, FMT_DEC2)
    ctx.advance()

    write_label(ctx, "B", "Margin of Safety", bold=True)
    write_formula(
        ctx,
        "C",
        f'=IF({fair_price}=0,0,({fair_price}-{current_price})/{fair_price})',
        FMT_PCT,
        is_key_total=True,
    )
    ctx.advance()

    write_label(ctx, "B", "Recommendation", bold=True)
    rec_cell = ctx.ws[f"C{ctx.row}"]
    rec_cell.value = f'=IF({fair_price}>{current_price},"Buy","Sell")'
    rec_cell.font = BLACK_BOLD
    rec_cell.fill = KEY_TOTAL_FILL
    rec_cell.alignment = CENTER_ALIGN
    rec_cell.number_format = "@"
    rec_cell.border = THIN_BORDER
    ctx.advance()


# ===== Top-level orchestration =====
def build(payload: dict, output_path: str) -> str:
    if "error" in payload:
        raise ValueError(f"MCP error for {payload.get('symbol', 'UNKNOWN')}: {payload['error']}")

    symbol = payload.get("symbol", "UNKNOWN").upper()
    dr = payload["discount_rate"]
    ge = payload["growth_estimates"]
    dt = payload["dcf_template"]
    dv = payload["dcf_value"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"DCF {symbol}"

    # Column widths (B=label, C=value, D..M projection years)
    widths = {
        "A": 2,
        "B": 38,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 18,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.showGridLines = False

    ctx = BuildContext(ws=ws, row=1)
    build_discount_rate_section(ctx, dr)
    build_growth_estimates_section(ctx, ge)
    build_dcf_template_section(ctx, dt)
    build_dcf_value_section(ctx, dv)

    output_path = os.path.abspath(output_path)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_json", help="Path to JSON file containing MCP get_stock_dcf_analysis output.")
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="Output .xlsx path. Defaults to ./{SYMBOL}_DCF.xlsx in cwd.",
    )
    args = parser.parse_args()

    with open(args.input_json, "r", encoding="utf-8") as f:
        payload = json.load(f)

    output_path = args.output_xlsx
    if output_path is None:
        symbol = (payload.get("symbol") or "DCF").upper()
        output_path = os.path.join(os.getcwd(), f"{symbol}_DCF.xlsx")

    written = build(payload, output_path)
    print(written)
    return 0


if __name__ == "__main__":
    sys.exit(main())
